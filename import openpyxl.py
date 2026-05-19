import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

# 제목 행
ws["A1"] = "측정 시간"
ws["B1"] = "체온"
ws["C1"] = "상태"

# 입력받아서 저장
for i in range(2, 5):
    시간 = input(str(i-1) + "번째 측정 시간: ")
    체온 = float(input("체온: "))
    
    if 체온 >= 38.0:
        상태 = "병원 가세요!"
    elif 체온 >= 37.5:
        상태 = "미열이에요"
    else:
        상태 = "정상이에요"
    
    ws.cell(row=i, column=1).value = 시간
    ws.cell(row=i, column=2).value = 체온
    ws.cell(row=i, column=3).value = 상태

wb.save("체온기록.xlsx")
print("엑셀 저장 완료!")